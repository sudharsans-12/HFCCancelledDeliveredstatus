"""
excel_writer.py
----------------
Writes the Cancelled and Delivered reports to styled .xlsx files:
  - Bold header row with fill
  - Green/Red fill on Shipping Carrier Status in the Cancelled report
  - Clickable hyperlinks in the Tracking Link column of the Delivered report
  - Auto-sized columns
"""

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(name="Arial", color="006100")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(name="Arial", color="9C0006")
REVIEW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
REVIEW_FONT = Font(name="Arial", color="9C6500")
HYPERLINK_FONT = Font(name="Arial", color="0563C1", underline="single")


def _write_dataframe(ws, df: pd.DataFrame):
    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")

    for row in df.itertuples(index=False):
        ws.append(list(row))

    for i, col in enumerate(df.columns, start=1):
        max_len = max(
            [len(str(col))] + [len(str(v)) for v in df[col].astype(str).tolist()]
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 10), 45)

    ws.freeze_panes = "A2"


def build_cancelled_workbook(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cancelled Report"
    _write_dataframe(ws, df)

    if "Carrier Status Color" in df.columns:
        col_idx = list(df.columns).index("Carrier Status Color") + 1
        color_map = {
            "GREEN": (GREEN_FILL, GREEN_FONT),
            "RED": (RED_FILL, RED_FONT),
            "REVIEW": (REVIEW_FILL, REVIEW_FONT),
        }
        for r, val in enumerate(df["Carrier Status Color"].tolist(), start=2):
            fill, font = color_map.get(val, (None, None))
            if fill:
                cell = ws.cell(row=r, column=col_idx)
                cell.fill = fill
                cell.font = font
    else:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = BODY_FONT

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_delivered_workbook(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Delivered Report"

    # Format date columns as strings first so openpyxl doesn't fight datetime dtypes
    df = df.copy()
    for date_col in ("Date", "Last Day Delivery"):
        if date_col in df.columns:
            df[date_col] = pd.to_datetime(df[date_col], errors="coerce").dt.strftime("%d-%m-%Y %H:%M")

    _write_dataframe(ws, df)

    if "Tracking Link" in df.columns:
        col_idx = list(df.columns).index("Tracking Link") + 1
        for r, url in enumerate(df["Tracking Link"].tolist(), start=2):
            if url:
                cell = ws.cell(row=r, column=col_idx)
                cell.hyperlink = url
                cell.font = HYPERLINK_FONT
                cell.value = url

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.font is None or cell.font.name != "Arial":
                cell.font = BODY_FONT

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
